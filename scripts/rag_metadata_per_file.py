#!/usr/bin/env python
import os, sys, time, re, requests
from pathlib import Path
from rag_fetch_files import read_csv, read_xlsx, read_edf  # tes helpers

MODEL = os.getenv("OLLAMA_MODEL", "qwen2.5:0.5b")
HOST  = os.getenv("OLLAMA_HOST", "127.0.0.1")
PORT  = os.getenv("OLLAMA_PORT", "11434")
NUM_CTX = int(os.getenv("NUM_CTX", "16384"))
NUM_PREDICT = int(os.getenv("NUM_PREDICT", "1500"))
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "900"))

PROMPT_TMPL = """You are a healthcare sleep data catalog assistant. Your job is to extract metadata fields (schema) exhaustively from the provided CONTEXT and produce a single Markdown table. Do not summarize events or measurements; list field names only with one concise definition each.

WHAT COUNTS AS METADATA
- CSV/XLSX: column headers, sheet names, and any header-like keys detected from the first rows (e.g., header-looking lines, key=value tokens, units in parentheses, multilingual labels).
- EDF header (global): version, patient_id, recording_id, startdate, starttime, header_bytes, n_records, record_duration_s, n_signals, etc. (ONLY if present in CONTEXT).
- EDF per-signal: label, transducer, phys_dim, phys_min, phys_max, dig_min, dig_max, prefilter, samp_per_record, fs_hz (ONLY if present).

EXHAUSTIVE HEADER DETECTION (CSV/XLSX)
Use ONLY what is in CONTEXT. If headers are missing/ambiguous, infer from the visible rows and mark example_data from the same lines:
1) If a header row exists (first non-empty row with multiple textual cells), treat it as headers.
2) If the first row is empty/merged, look at the next rows for likely headers (mostly text, distinct from numeric patterns).
3) If there are two-tier headers (e.g., "AHI (index)" on row1 and a sublabel row2), combine into one normalized name (e.g., apnea_hypopnea_index).
4) If headers are repeated or multilingual, deduplicate and keep one normalized name.
5) If no explicit headers exist, infer candidates by column position using the first 5–10 rows:
   - Prefer tokens that repeat in a column (e.g., S0001, S0002 → folder_id).
   - Detect key=value patterns and lift the keys as metadata fields.
   - Detect units/labels in parentheses (e.g., "SpO2 (%)" → spo2_percent).
   - Use domain hints (AHI, ODI, PLMI, AI, HI, LMI, start_time, study_type, scorer) ONLY if those tokens appear in CONTEXT.
6) Sheet-level metadata (XLSX): if sheet names are visible (e.g., "TME", "SIM"), include a field sheet_name.
7) Normalize field names to lower_snake_case (e.g., "Apnea-Hypopnea Index" → apnea_hypopnea_index; "Start Time" → start_time).
8) Never invent fields not evidenced by the CONTEXT. Every row must be justified by a visible header/token/pattern in the CONTEXT.

OUTPUT CONSTRAINTS (STRICT)
- Output ONLY the table (no prose, no explanations).
- Table format (exactly these 4 columns, in this order):

| source | metadata_field | example_data | proposed_definition |

- source ∈ {{csv,xlsx,edf_header,edf_signal}}.
- metadata_field: normalized lower_snake_case.
- example_data: one short representative value seen in CONTEXT for that field, otherwise "—".
- proposed_definition: concise, technical definition; include unit if visible (e.g., "Apnea–Hypopnea Index (events/hour)").
- Be exhaustive: include ALL detectable fields from the CONTEXT (headers, key-like tokens, EDF header keys, EDF signal attributes).

HALLUCINATION GUARDRAILS
- Use ONLY strings/tokens present in CONTEXT.
- Do NOT output timelines or per-epoch events (start/end/duration lines) as metadata.
- If unsure about a field name but there is a consistent column pattern, propose a conservative normalized name and set example_data from CONTEXT; otherwise skip.

---
CONTEXT:
{context}
---
QUESTION:
Build ONLY the required Markdown table (no other text).
"""

def wait_for_server(timeout=30):
    url = f"http://{HOST}:{PORT}/api/tags"
    t0 = time.time()
    while time.time() - t0 < timeout:
        try:
            r = requests.get(url, timeout=5)
            if r.ok: return True
        except requests.exceptions.RequestException:
            pass
        time.sleep(1)
    return False

def ask(prompt: str, retries=3, backoff=5) -> str:
    if not wait_for_server(): raise RuntimeError("Ollama API not ready.")
    url = f"http://{HOST}:{PORT}/api/generate"
    payload = {
        "model": MODEL,
        "prompt": prompt,
        "stream": False,
        "options": {
            "num_ctx": NUM_CTX,
            "num_predict": NUM_PREDICT,
            "temperature": 0.1,
            "top_p": 0.9,
            "repeat_penalty": 1.15,
        },
    }
    last = None
    for i in range(retries + 1):
        try:
            r = requests.post(url, json=payload, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return r.json().get("response","").strip()
        except requests.exceptions.RequestException as e:
            last = e
            if i < retries: time.sleep(backoff * (2**i)); continue
            raise last

def file_to_context(p: Path, max_rows: int = 300) -> str:
    ext = p.suffix.lower()
    # Réduire au strict nécessaire (headers/meta) pour éviter de dépasser le contexte
    if ext == ".csv":
        import csv
        with open(p, "r", encoding="utf-8", errors="ignore", newline="") as f:
            r = csv.reader(f)
            headers = next(r, [])
        return "file={name} kind=csv_header | columns={cols}".format(
            name=p.name, cols=",".join([str(h).strip() for h in headers if h is not None])
        )
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        headers = next(ws.iter_rows(values_only=True), [])
        wb.close()
        return "file={name} kind=xlsx_header | columns={cols}".format(
            name=p.name, cols=",".join([str(h).strip() for h in headers if h is not None])
        )
    if ext == ".edf":
        # 1 ligne header + au plus 20 signaux
        lines = read_edf(p, max_rows=20)  # ta read_edf renvoie déjà une liste de lignes
        return "\n".join(lines)
    return ""

def sanitize_filename(name: str) -> str:
    # remplace espaces/accents/symboles pour un nom de fichier sûr
    base = re.sub(r"[^\w\.-]+","_", name, flags=re.UNICODE)
    return base.strip("_") or "untitled"

def write_table_md(out_file: Path, response_text: str):
    # Ne garder que le tableau; header inclus
    lines = [ln for ln in response_text.splitlines() if ln.strip()]
    try:
        start = next(i for i,l in enumerate(lines) if l.strip().startswith("| source |"))
    except StopIteration:
        return False
    table = lines[start:]
    out_file.parent.mkdir(parents=True, exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        f.write("\n".join(table) + "\n")
        f.flush(); os.fsync(f.fileno())
    return True

def main(root="data/raw", out_dir="outputs", max_rows=200):
    root = Path(root); out_root = Path(out_dir)
    files = [p for p in sorted(root.rglob("*.*")) if p.suffix.lower() in (".csv",".xlsx",".edf")]
    total = len(files)
    for idx, p in enumerate(files, start=1):
        print(f"[{idx}/{total}] {p}", flush=True)
        try:
            ctx = file_to_context(p, max_rows)
            if not ctx.strip(): 
                print(f"# skip (empty context) {p}", file=sys.stderr); continue
            prompt = PROMPT_TMPL.format(context=ctx, question="Build ONLY the required Markdown table (no other text).")
            resp = ask(prompt)
            out_file = out_root / f"meta_{sanitize_filename(p.name)}.md"
            ok = write_table_md(out_file, resp)
            if not ok:
                print(f"# skip (no table) {p}", file=sys.stderr)
        except Exception as e:
            print(f"# skip {p}: {e}", file=sys.stderr)
        time.sleep(1)  # laisse souffler l'API entre fichiers

if __name__ == "__main__":
    root   = sys.argv[1] if len(sys.argv) > 1 else "data/raw"
    outdir = sys.argv[2] if len(sys.argv) > 2 else "outputs"
    maxr   = int(sys.argv[3]) if len(sys.argv) > 3 else 200
    main(root, outdir, maxr)
