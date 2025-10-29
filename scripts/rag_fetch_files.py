#!/usr/bin/env python
import os, sys, csv, argparse, struct
from pathlib import Path
from openpyxl import load_workbook

# ---------------------------
# Helpers CSV / XLSX
# ---------------------------
def read_csv(path: Path, max_rows: int):
    out = []
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        sample = f.read(4096); f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        headers = next(reader, [])
        for i, row in enumerate(reader, start=1):
            if max_rows and i > max_rows: break
            pairs = [f"{h}={v}" for h, v in zip(headers, row)]
            out.append(f"file={path.name} row={i} | " + " | ".join(pairs))
    return out

def read_xlsx(path: Path, max_rows: int):
    out = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, [])
    for i, row in enumerate(rows, start=1):
        if max_rows and i > max_rows: break
        pairs = [f"{str(h)}={str(v)}" for h, v in zip(headers, row)]
        out.append(f"file={path.name} row={i} | " + " | ".join(pairs))
    wb.close()
    return out

# ---------------------------
# EDF header reader (light)
# Lit uniquement les métadonnées (pas les signaux bruts)
# ---------------------------
def _read_str(b: bytes) -> str:
    return b.decode("ascii", errors="ignore").strip()

def read_edf_header(path: Path):
    """
    Retourne (header_dict, signals_list) où:
      header_dict = { 'version', 'patient_id', 'recording_id', 'startdate', 'starttime',
                      'header_bytes', 'n_records', 'record_duration_s', 'n_signals' }
      signals_list = [ { 'label','transducer','phys_dim','phys_min','phys_max',
                          'dig_min','dig_max','prefilter','samp_per_record','fs_hz' }, ... ]
    """
    with open(path, "rb") as f:
        H = f.read(256)
        if len(H) < 256:
            raise ValueError("EDF header too small")
        version        = _read_str(H[0:8])
        patient_id     = _read_str(H[8:88])
        recording_id   = _read_str(H[88:168])
        startdate      = _read_str(H[168:176])  # dd.mm.yy
        starttime      = _read_str(H[176:184])  # hh.mm.ss
        header_bytes   = int(_read_str(H[184:192]) or 0)
        reserved       = _read_str(H[192:236])
        n_records      = int(float(_read_str(H[236:244]) or 0))
        record_dur_sec = float(_read_str(H[244:252]) or 0.0)
        n_signals      = int(_read_str(H[252:256]) or 0)

        # Chaque bloc de metadata par signal est contigu (taille fixe)
        def read_signal_strings(field_len, nsig):
            return [_read_str(f.read(field_len)) for _ in range(nsig)]

        labels        = read_signal_strings(16,  n_signals)
        transducers   = read_signal_strings(80,  n_signals)
        phys_dims     = read_signal_strings(8,   n_signals)
        phys_mins     = read_signal_strings(8,   n_signals)
        phys_maxs     = read_signal_strings(8,   n_signals)
        dig_mins      = read_signal_strings(8,   n_signals)
        dig_maxs      = read_signal_strings(8,   n_signals)
        prefilters    = read_signal_strings(80,  n_signals)
        samp_per_rec  = read_signal_strings(8,   n_signals)
        _reserved2    = read_signal_strings(32,  n_signals)

    header = {
        "version": version,
        "patient_id": patient_id,
        "recording_id": recording_id,
        "startdate": startdate,
        "starttime": starttime,
        "header_bytes": header_bytes,
        "n_records": n_records,
        "record_duration_s": record_dur_sec,
        "n_signals": n_signals,
    }

    signals = []
    for i in range(n_signals):
        try:
            n_samp = int(float(samp_per_rec[i] or 0))
        except Exception:
            n_samp = 0
        fs = (n_samp / record_dur_sec) if record_dur_sec else 0.0
        sig = {
            "label": labels[i],
            "transducer": transducers[i],
            "phys_dim": phys_dims[i],
            "phys_min": phys_mins[i],
            "phys_max": phys_maxs[i],
            "dig_min": dig_mins[i],
            "dig_max": dig_maxs[i],
            "prefilter": prefilters[i],
            "samp_per_record": n_samp,
            "fs_hz": round(fs, 6),
        }
        signals.append(sig)

    return header, signals

def read_edf(path: Path, max_rows: int):
    """
    Émet des lignes texte pour RAG :
      - 1 ligne header
      - jusqu’à max_rows lignes de signaux (label, fs, etc.)
    """
    header, signals = read_edf_header(path)
    out = []
    out.append(
        "file={name} kind=edf_header | patient_id={pid} | recording_id={rid} | "
        "start={date}T{time} | n_records={nr} | record_dur_s={dur} | n_signals={ns}".format(
            name=path.name, pid=header["patient_id"], rid=header["recording_id"],
            date=header["startdate"], time=header["starttime"],
            nr=header["n_records"], dur=header["record_duration_s"], ns=header["n_signals"]
        )
    )
    limit = max_rows if max_rows else len(signals)
    for i, s in enumerate(signals[:limit], start=1):
        out.append(
            "file={name} kind=edf_signal row={i} | label={lab} | fs_hz={fs} | phys_dim={pd} | "
            "phys_min={pmin} | phys_max={pmax} | dig_min={dmin} | dig_max={dmax} | prefilter={pf}".format(
                name=path.name, i=i, lab=s["label"], fs=s["fs_hz"], pd=s["phys_dim"],
                pmin=s["phys_min"], pmax=s["phys_max"], dmin=s["dig_min"], dmax=s["dig_max"],
                pf=s["prefilter"]
            )
        )
    return out

# ---------------------------
# Collecteur générique
# ---------------------------
def collect(root: Path, max_rows_per_file: int, glob: str):
    lines = []
    files = sorted(root.rglob(glob))
    for p in files:
        try:
            ext = p.suffix.lower()
            if ext == ".csv":
                lines += read_csv(p, max_rows_per_file)
            elif ext in (".xlsx",):
                lines += read_xlsx(p, max_rows_per_file)
            elif ext in (".edf",):
                lines += read_edf(p, max_rows_per_file)
        except Exception as e:
            print(f"# skip {p}: {e}", file=sys.stderr)
    return "\n".join(lines)

# ---------------------------
# CLI
# ---------------------------
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="data/raw", help="répertoire racine des données")
    ap.add_argument("--glob", default="**/*.*", help="pattern (ex: '**/*.csv')")
    ap.add_argument("--max-rows-per-file", type=int, default=500, help="limite de lignes lues par fichier")
    args = ap.parse_args()
    print(collect(Path(args.root), args.max_rows_per_file, args.glob))
